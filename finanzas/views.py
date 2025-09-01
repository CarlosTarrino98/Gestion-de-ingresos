from django.shortcuts import render, redirect, get_object_or_404
from .models import Ingreso, Gasto
from .forms import IngresoForm, GastoForm
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from openpyxl.styles import Font
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

@login_required
def lista_ingresos(request):
    ingresos = Ingreso.objects.filter(user=request.user).order_by('-fecha')

    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')

    if fecha_desde:
        ingresos = ingresos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        ingresos = ingresos.filter(fecha__lte=fecha_hasta)

    return render(request, 'finanzas/lista_ingresos.html', {
        'ingresos': ingresos,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })

@login_required
def lista_gastos(request):
    gastos = Gasto.objects.filter(user=request.user).order_by('-fecha')

    fecha_desde = request.GET.get('desde')
    fecha_hasta = request.GET.get('hasta')

    if fecha_desde:
        gastos = gastos.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        gastos = gastos.filter(fecha__lte=fecha_hasta)

    return render(request, 'finanzas/lista_gastos.html', {
        'gastos': gastos,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    })

@login_required
def añadir_ingreso(request):
    if request.method == 'POST':
        form = IngresoForm(request.POST)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.user = request.user
            ingreso.save()
            return redirect('lista_ingresos')
    else:
        form = IngresoForm()
    return render(request, 'finanzas/form_ingreso.html', {'form': form})

@login_required
def añadir_gasto(request):
    if request.method == 'POST':
        form = GastoForm(request.POST)
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.user = request.user
            gasto.save()
            return redirect('lista_gastos')
    else:
        form = GastoForm()
    return render(request, 'finanzas/form_gasto.html', {'form': form})

@login_required
def editar_ingreso(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk, user=request.user)
    if request.method == 'POST':
        form = IngresoForm(request.POST, instance=ingreso)
        if form.is_valid():
            ingreso = form.save(commit=False)
            ingreso.user = request.user
            ingreso.save()
            return redirect('lista_ingresos')
    else:
        form = IngresoForm(instance=ingreso)
    return render(request, 'finanzas/form_ingreso.html', {'form': form})

@login_required
def editar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk, user=request.user)
    if request.method == 'POST':
        form = GastoForm(request.POST, instance=gasto)
        if form.is_valid():
            gasto = form.save(commit=False)
            gasto.user = request.user
            gasto.save()
            return redirect('lista_gastos')
    else:
        form = GastoForm(instance=gasto)
    return render(request, 'finanzas/form_gasto.html', {'form': form})

@login_required
def eliminar_ingreso(request, pk):
    ingreso = get_object_or_404(Ingreso, pk=pk, user=request.user)
    ingreso.delete()
    return redirect('lista_ingresos')

@login_required
def eliminar_gasto(request, pk):
    gasto = get_object_or_404(Gasto, pk=pk, user=request.user)
    gasto.delete()
    return redirect('lista_gastos')

@login_required
def generar_seguimiento(request):
    fecha_inicio = parse_date(request.GET.get('fecha_inicio'))
    fecha_fin = parse_date(request.GET.get('fecha_fin'))
    formato = request.GET.get('formato')  # 'excel' o 'pdf'

    # Asegurarse de que sean objetos date
    if isinstance(fecha_inicio, str):
        fecha_inicio = parse_date(fecha_inicio)
    if isinstance(fecha_fin, str):
        fecha_fin = parse_date(fecha_fin)

    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Fechas inválidas", status=400)

    ingresos = Ingreso.objects.filter(user=request.user, fecha__range=(fecha_inicio, fecha_fin))
    gastos = Gasto.objects.filter(user=request.user, fecha__range=(fecha_inicio, fecha_fin))

    total_ingresos = sum(i.cantidad for i in ingresos)
    total_gastos = sum(g.cantidad for g in gastos)
    total_saldo = total_ingresos - total_gastos

    # Formatear fechas para el nombre del archivo
    fecha_ini_str = fecha_inicio.strftime('%d-%m-%Y')
    fecha_fin_str = fecha_fin.strftime('%d-%m-%Y')
    nombre_archivo = f"ODA - {fecha_ini_str} - {fecha_fin_str}"

    if formato == 'excel':
        wb = Workbook()
        ws_ingresos = wb.active
        ws_ingresos.title = "Ingresos"

        # Estilos
        cabecera_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        def estilizar_fila(celdas):
            for cell in celdas:
                cell.font = bold_font
                cell.fill = cabecera_fill
                cell.alignment = center_align
                cell.border = thin_border

        # Hoja Ingresos
        ws_ingresos.append(["Fecha", "Concepto", "Cantidad"])
        estilizar_fila(ws_ingresos[1])
        for ingreso in ingresos:
            ws_ingresos.append([
                ingreso.fecha.strftime('%d/%m/%Y'),
                ingreso.concepto,
                ingreso.cantidad
            ])
        for row in ws_ingresos.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border

        # Hoja Gastos
        ws_gastos = wb.create_sheet("Gastos")
        ws_gastos.append(["Fecha", "Concepto", "Cantidad"])
        estilizar_fila(ws_gastos[1])
        for gasto in gastos:
            ws_gastos.append([
                gasto.fecha.strftime('%d/%m/%Y'),
                gasto.concepto,
                gasto.cantidad
            ])
        for row in ws_gastos.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border

        # Hoja Totales
        ws_totales = wb.create_sheet("Totales")
        ws_totales.append(["Resumen", "Cantidad (€)"])
        estilizar_fila(ws_totales[1])
        ws_totales.append(["Total Ingresos", total_ingresos])
        ws_totales.append(["Total Gastos", total_gastos])
        ws_totales.append(["Saldo", total_saldo])
        for row in ws_totales.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename={nombre_archivo}.xlsx'
        wb.save(response)
        return response

    elif formato == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'

        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("ODA - Resumen de seguimiento", styles['Title']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<b>Desde:</b> {fecha_inicio.strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Paragraph(f"<b>Hasta:</b> {fecha_fin.strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 12))
        elements.append(Paragraph(f"<b>Total Ingresos:</b> {total_ingresos:.2f} €", styles['Normal']))
        elements.append(Paragraph(f"<b>Total Gastos:</b> {total_gastos:.2f} €", styles['Normal']))
        elements.append(Paragraph(f"<b>Saldo:</b> {total_saldo:.2f} €", styles['Normal']))
        elements.append(Spacer(1, 24))

        # Tabla de ingresos
        elements.append(Paragraph("Listado de Ingresos", styles['Heading2']))
        data_ingresos = [["Fecha", "Concepto", "Cantidad (€)"]]
        for ingreso in ingresos:
            data_ingresos.append([
                ingreso.fecha.strftime('%d/%m/%Y'),
                Paragraph(ingreso.concepto, styles['Normal']),
                f"{ingreso.cantidad:.2f}"
            ])
        table_ingresos = Table(data_ingresos, colWidths=[80, 300, 80], hAlign='LEFT')
        table_ingresos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        elements.append(table_ingresos)
        elements.append(Spacer(1, 24))

        # Tabla de gastos
        elements.append(Paragraph("Listado de Gastos", styles['Heading2']))
        data_gastos = [["Fecha", "Concepto", "Cantidad (€)"]]
        for gasto in gastos:
            data_gastos.append([
                gasto.fecha.strftime('%d/%m/%Y'),
                Paragraph(gasto.concepto, styles['Normal']),
                f"{gasto.cantidad:.2f}"
            ])
        table_gastos = Table(data_gastos, colWidths=[80, 300, 80], hAlign='LEFT')
        table_gastos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.salmon),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP')
        ]))
        elements.append(table_gastos)

        doc.build(elements)
        return response

    return HttpResponse("Formato no válido", status=400)